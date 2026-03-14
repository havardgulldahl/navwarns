"""
Script to extract NAVAREA XIX history from an Excel file from Kystverket, and
convert it to NavwarnMessage objects via the shared parser.
"""

import argparse
import datetime
import json
import logging
import re
import sys
from typing import Dict, List, Optional, Set, Tuple
import pathlib

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
import openpyxl
from datetime import datetime

this_dir = pathlib.Path(__file__).resolve().parent
try:
    from . import parser as navparser  # type: ignore
    from . import cleanup  # type: ignore
    from . import scraper
except ImportError:  # running as a script
    import importlib.util

    parser_path = this_dir / "parser.py"
    spec = importlib.util.spec_from_file_location("navparser", parser_path)
    navparser = importlib.util.module_from_spec(spec)  # type: ignore
    assert spec and spec.loader
    spec.loader.exec_module(navparser)  # type: ignore

    cleanup_path = this_dir / "cleanup.py"
    spec_clean = importlib.util.spec_from_file_location("cleanup", cleanup_path)
    cleanup = importlib.util.module_from_spec(spec_clean)  # type: ignore
    assert spec_clean and spec_clean.loader
    spec_clean.loader.exec_module(cleanup)  # type: ignore

    scraper_path = this_dir / "scraper.py"
    spec_scraper = importlib.util.spec_from_file_location("scraper", scraper_path)
    scraper = importlib.util.module_from_spec(spec_scraper)  # type: ignore
    assert spec_scraper and spec_scraper.loader
    spec_scraper.loader.exec_module(scraper)  # type: ignore


# Path to the Excel file
file_path = this_dir.parent / "history" / "NAVAREA XIX varsler frem til 14.03.2026.xlsx"

CURRENT_DIR = pathlib.Path("current")
HISTORY_DIR = this_dir.parent / "history"


def extract_warnings(excel_file: pathlib.Path) -> List[Dict[str, str]]:
    # Load the workbook and select the first sheet
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    # Extract headers
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    # Find relevant column indices
    number_idx = headers.index("Beskrivelse")
    raw_navwarn_idx = headers.index("KommentarEN")
    start_timestamp_idx = headers.index("PlanlagtDato")
    end_timestamp_idx = headers.index("AvsluttetDato")

    # Collect history entries
    history = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_navwarn = row[raw_navwarn_idx]
        try:
            number = str(row[number_idx]).split(":")[1].strip()
        except IndexError:
            # get first line of raw_navwarn as fallback
            number = str(raw_navwarn).splitlines()[0].strip()
        # dates are in Excel datetime format, but openpyxl should have converted
        # them to Python datetimes already.
        start_timestamp: datetime = (
            row[start_timestamp_idx]
            if isinstance(row[start_timestamp_idx], datetime)
            else None
        )
        end_timestamp: datetime = (
            row[end_timestamp_idx]
            if isinstance(row[end_timestamp_idx], datetime)
            else None
        )
        history.append(
            {
                "number": number,
                "raw_navwarn": raw_navwarn,
                "start_timestamp": start_timestamp,
                "end_timestamp": end_timestamp,
            }
        )
    logging.info("Extracted %d entries from Excel", len(history))
    return history


def main(dry_run: bool = True, overwrite_files: bool = False):
    logging.info("Extracting NAVAREA XIX history from Excel file: %s", file_path)

    warnings = extract_warnings(file_path)
    if not warnings:
        logging.info("No Historic NAVAREA XIX warnings extracted.")
        return

    all_stored: Set[str] = set()
    for warn in warnings:
        body = warn["raw_navwarn"]

        # Parse as standard NAVWARN message
        coords = navparser.parse_coordinates(body)
        geometry, radius = navparser.analyze_geometry(body, coords)

        navmsg = navparser.NavwarnMessage(
            msg_id=f"NAVAREA XIX {warn['number']}",
            dtg=warn["start_timestamp"],
            raw_dtg="",
            cancel_date=(
                warn["end_timestamp"].isoformat() if warn["end_timestamp"] else None
            ),
            year=warn["start_timestamp"].year if warn["start_timestamp"] else None,
            body=warn["raw_navwarn"],
            coordinates=coords,
            geometry=geometry,
            radius=radius,
            cancellations=navparser.parse_cancellations(body),
            hazard_type=navparser.classify_hazard(body),
            groups=navparser.parse_coordinate_groups(body),
        )

        if dry_run:
            print(navmsg.to_geojson_features())

        else:

            # go through all, selecting by year
            stored = scraper.store_messages(
                [
                    navmsg,
                ],
                force=overwrite_files,
                output_dir=HISTORY_DIR / str(warn["start_timestamp"].year) / "navwarns",
            )
            all_stored.update(stored)

    logging.info(
        f"Processed and saved {len(warnings)} messages into {len(all_stored)} files (overwrite={overwrite_files})."
    )


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Extract history XIX from excel")
    ap.add_argument(
        "--force", action="store_true", help="Overwrite existing message JSON files"
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Print parsed messages to stdout instead of writing files",
    )
    args = ap.parse_args()
    main(dry_run=args.dry_run, overwrite_files=args.force)
